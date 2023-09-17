import openpyxl

# Load the Excel workbook
workbook = openpyxl.load_workbook('driver/Account.xlsx')  # Replace with the actual file path

# Select the specific sheet you want to work with
sheet = workbook.active  # Use sheet = workbook['SheetName'] for a named sheet

# Find the column containing email addresses (assuming it's in the first column)
email_column = sheet['A']  # Change 'A' to the appropriate column letter

# Iterate through the rows, assuming the email addresses are in column A
for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=1):
    email = row[0].value
    print("Working with: " + email)
    sheet.cell(row=row[0].row, column=row[0].column + 1, value=email)

# Save the modified workbook
workbook.save('your_updated_excel_file.xlsx')  # Replace with the desired output file path

# Close the workbook
workbook.close()
