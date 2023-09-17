from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.support.wait import WebDriverWait
import openpyxl

# set mobile user agent setting
mobile_emulation = {
    "deviceMetrics": {"width": 360, "height": 640, "pixelRatio": 3.0},
    "userAgent": "Mozilla/5.0 (Linux; Android 10; SM-A205F) AppleWebKit/537.36 (KHTML, like Gecko) "
                 "Chrome/90.0.4430.210 Mobile Safari/537.36",
}
chrome_options = webdriver.ChromeOptions()
chrome_options.add_experimental_option("mobileEmulation", mobile_emulation)

# set chrome driver
service = Service('driver/chromedriver.exe')
driver = webdriver.Chrome(service=service, options=chrome_options)
driver.implicitly_wait(10)
driver.maximize_window()
wait = WebDriverWait(driver, 10)

# Load the Excel workbook
workbook = openpyxl.load_workbook('driver/Daraz.xlsx')
sheet = workbook.active
for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=1):
    try:
        email = str(row[0].value)
        print("Working with: " + email)
        driver.get('https://member-m.daraz.com.bd/user/login')
        (driver.find_element(By.CSS_SELECTOR, 'input[type="text"]').send_keys(email))
        driver.find_element(By.CSS_SELECTOR, "input[type='password']").send_keys('qusal.Sm34')
        wait.until(ec.element_to_be_clickable((By.CLASS_NAME, 'login-btn'))).click()
        wait.until(ec.element_to_be_clickable((By.ID, 'my-orders'))).click()
        wait.until(ec.element_to_be_clickable((By.XPATH, "//span[contains(text(),'Got It')]"))).click()
        driver.find_element(By.XPATH, "//span[contains(text(),'>')]").click()
        tracking_element = driver.find_element(By.CSS_SELECTOR, "div[t='package-operations'] :nth-child(1)")
        driver.execute_script('arguments[0].scrollIntoView();', tracking_element)
        wait.until(
            ec.element_to_be_clickable(tracking_element)).click()
        tracking_number = wait.until(
            ec.element_to_be_clickable((By.CSS_SELECTOR, '[id^="DEX-BDN"] span'))).text
        print(tracking_number)
        sheet.cell(row=row[0].row, column=row[0].column + 2, value=tracking_number)
        address = wait.until(ec.element_to_be_clickable((By.XPATH, "//body/div[@id='root_root']/div[1]/div["
                                                                   "3]/div[1]/div[1]/div[2]/span"))).text
        print(address)
        sheet.cell(row=row[0].row, column=row[0].column + 1, value=address)
        try:
            otp = wait.until(ec.element_to_be_clickable((By.CSS_SELECTOR, "[id='root_root'] div:nth-child(1) "
                                                                          "div:nth-child(4) div"))).text
        except Exception as e:
            print(e)
            otp = ""
            # keep only the digits
            otp = ''.join(filter(str.isdigit, otp))
            print(otp)

        # keep only the digits
        otp = ''.join(filter(str.isdigit, otp))
        print(otp)
        sheet.cell(row=row[0].row, column=row[0].column + 3, value=otp)
        # clear browser
        driver.delete_all_cookies()
        # Save the modified workbook
        workbook.save('driver/updated_excel_file.xlsx')
    except Exception as e:
        print(e)
        input()
        driver.delete_all_cookies()

workbook.close()
# prevent browser from closing
input('Press ENTER to close the automated browser')
driver.quit()
